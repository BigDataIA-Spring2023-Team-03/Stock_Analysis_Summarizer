import streamlit as st
import pandas as pd
import openpyxl
# import os
# import requests

# Adding a wallpaper to the web application
def add_bg_from_url():
    st.markdown(
         f"""
         <style>
         .stApp {{
             background-image: url("https://static.seekingalpha.com/cdn/s3/uploads/getty_images/1057996634/image_1057996634.jpg?io=getty-c-w1280.jpg");
             background-attachment: fixed;
             background-size: cover
         }}
         </style>
         """,
         unsafe_allow_html=True
     )

add_bg_from_url() 

# Title of the page
st.title("Stock Analysis Summarizer")

# Uploading User Stock Portfolio
file = st.file_uploader("**Upload your Stock Portfolio**", type=["xlsx"])

# If the user has uploaded a file
if file is not None:
    # Read the Excel file into a Pandas DataFrame
    df = pd.read_excel(file, sheet_name='Sheet1')
    # Display the contents of the DataFrame
    st.write(df)

# Selecting Ticker for Analysis from the List Available
add_item = st.text_input("**Add Ticker**")
if add_item:
    # Load the Excel file
    workbook = openpyxl.load_workbook("C:/Users/rumij/Downloads/ticker.xlsx")
    # Select the worksheet you want to add the item to
    worksheet = workbook["Sheet1"]
    # Add a new item to the worksheet
    new_item = add_item
    worksheet.append([new_item])
    # Save the changes to the Excel file
    workbook.save("C:/Users/rumij/Downloads/ticker.xlsx")

remove_item = st.text_input("**Delete a Ticker**")
if remove_item:
    # Load the Excel file
    workbook = openpyxl.load_workbook("C:/Users/rumij/Downloads/ticker.xlsx")
    # Select the worksheet you want to add the item to
    worksheet = workbook["Sheet1"]
    # Add a new item to the worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == remove_item:
            #     # Delete the row containing the cell
                worksheet.delete_rows(cell.row)
                # Save the changes to the Excel file
                workbook.save("C:/Users/rumij/Downloads/ticker.xlsx")
                break  # Exit the inner loop after deleting the row

my_list = []
stocks = pd.read_excel('C:/Users/rumij/Downloads/ticker.xlsx')
stocks = stocks['Ticker'].unique()
# selected_option = st.selectbox("**Select an option**", stocks)
selected_option = st.multiselect("**Select Any 5 Tickers for Analysis**", stocks, max_selections=2)

# my_list.append(selected_option)
my_list = 'You selected: ' + ', '.join(selected_option)

st.write(my_list)