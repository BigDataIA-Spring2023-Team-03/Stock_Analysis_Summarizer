import streamlit as st
import pandas as pd
import openpyxl
import yfinance as yf
import requests
import time
from datetime import datetime
# import snowflake.connector
from decouple import config
import boto3
import json

# Summarization
from transformers import pipeline, BartTokenizer

from Util import db_util

# DEV or PROD
environment = 'DEV'
if environment == 'DEV':
    fastapi_webserver = 'localhost:8000'
    airflow_webserver = 'localhost:8080'
elif environment == 'PROD':
    fastapi_webserver = 'backend:8000'
    airflow_webserver = '107.20.116.176:8080'

# AWS KEYS
aws_access_key_id = config('aws_access_key_id')
aws_secret_access_key = config('aws_secret_access_key')

# OpenAI
openai_api_key = config('openai_api_key')

# S3 Details:
s3_bucket_name = 'stock-analysis-summarizer'

# Create an S3 client
s3_client = boto3.client(
    "s3",
    aws_access_key_id=aws_access_key_id,
    aws_secret_access_key=aws_secret_access_key,
)


##########################################################################

# Function to ask a companies ticker symbol
def get_ticker(company_name):
    url = "https://api.openai.com/v1/engines/text-davinci-002/completions"
    Ticker = {
        "prompt": "Q: What is the ticker for" + " " + company_name + "?",
        "temperature": 0.7,
        "max_tokens": 50,
        "stop": None
    }

    # set up the API headers
    headers = {
        "Content-Type": "application/json",
        "Authorization": f'Bearer {openai_api_key}'  # "Bearer YOUR_KEY"
    }

    # send the API request
    response = requests.post(url, json=Ticker, headers=headers)

    # Exception Handeling
    if response.status_code == 200:
        data = response.json()
        generated_text = data["choices"][0]["text"]
        ticker_symbol = generated_text.split()
        st.write(ticker_symbol[-1])
    else:
        st.write("Error:", response.text)


# Function to validate the ticker symbol
def validate_ticker(ticker):
    try:
        data = yf.Ticker(ticker)
        info = data.info
    except:
        return False
    return True


# Function to add ticker manually
def add_data(ticker):
    st.session_state['portfolio'] = st.session_state['portfolio'].append({'Stock_Ticker': ticker}, ignore_index=True)


# Function to grab all the data for a stock from S3
def get_analysis_data(stock):
    analysis_dataset = []
    # TESTING
    # st.write('Analysis_Results/{stock}')
    folder = s3_client.list_objects_v2(Bucket=s3_bucket_name, Prefix=f'Analysis_Results/{stock}')
    print(folder)
    print(f"""Total Files in S3: {folder['KeyCount']}""")
    # If there are articles
    if folder['KeyCount'] > 0:
        for obj in folder['Contents']:
            files = obj['Key']
            print(files)
            # Pull JSON data from each file
            response = s3_client.get_object(Bucket=s3_bucket_name, Key=files)
            content = response['Body']
            analysis_dataset.extend(json.loads(content.read()))

        return pd.DataFrame(analysis_dataset, columns=['publish_date', 'bart_summary', 'sentiment'])
    else:
        return pd.DataFrame()


# Check DAG Status
def check_dag_status(dag_id):
    url = f'http://{airflow_webserver}/api/v1/dags/{dag_id}/dagRuns'
    response = requests.get(url=url, auth=('airflow2', 'airflow2'))
    response_json = response.json()
    state = response_json['dag_runs'][len(response_json['dag_runs']) - 1]['state']
    return state


# Get Summary of Content using Facebook BART Large CNN transformer model
def article_summary(sentiment, aggregate_summary):
    # Token text input for the Bart model, max 1024 tokens
    # tokenizer = BartTokenizer.from_pretrained('facebook/bart-large')

    # max_length = 1024
    # tokens = tokenizer(aggregate_summary, max_length=max_length, truncation=True, return_tensors='pt')

    # Summarize
    summarizer = pipeline("summarization", model="facebook/bart-large-cnn")

    summary = summarizer(aggregate_summary[:1024], max_length=130, min_length=30, do_sample=False)

    summary = summary[0]['summary_text']

    print(f'Summary created for {sentiment} articles')

    return summary


##############################################################################################################################

def portfolio_uploader():
    # Title of the page
    st.title("Stock Analysis Summarizer")

    # TODO: ADD BACK IN if there is time
    # Uploading User Stock Portfolio
    # excel_upload = st.checkbox('Upload Portfolio from File')
    # if excel_upload:
    #     file = st.file_uploader("**Upload your Stock Portfolio**", type=["xlsx"])

    # Initialize
    if "portfolio" not in st.session_state:
        st.session_state['portfolio'] = pd.DataFrame(columns=['Stock_Ticker'])

    # Clear portfolio
    clear = st.checkbox('Clear Portfolio')
    if clear:
        st.session_state['portfolio'] = pd.DataFrame(columns=['Stock_Ticker'])

    # Manually add tickers
    manual_upload = st.checkbox('Manually Upload Portfolio')

    if manual_upload:
        # Get companies ticker symbol from ChatGPT
        company_name = st.text_input("Enter Company Name for Ticker Symbol")
        if company_name:
            get_ticker(company_name)

        # Adding ticker manually
        portfolio = pd.DataFrame(columns=['Stock_Ticker'])

        # Only can run 5 stocks at a time
        if st.session_state['portfolio'].shape[0] <= 4:

            ticker = st.text_input("Enter a ticker:")

            if ticker:
                # Check if ticker already listed 
                if ticker not in st.session_state['portfolio']['Stock_Ticker'].tolist():
                    # Add Validation check for the ticker from Yahoo finanace
                    if validate_ticker(ticker):
                        st.write(f"{ticker} is a valid ticker symbol.")
                        # Adding the ticker manually
                        add_data(ticker)
                        st.session_state['portfolio'] = st.session_state['portfolio']
                    else:
                        st.error(f"{ticker} is not a valid ticker symbol.")
                else:
                    # TODO: RUN analysis is checked, then don't show error
                    # if run_analysis:
                    #     pass
                    # else:
                    st.error(f"{ticker} already listed.")
        else:
            st.error('You can only run 5 stocks at a time!')

        # TODO: Doesn't Work
        # Remove a ticker from the portfolio
        # for i, ticker in enumerate(st.session_state['portfolio']['Stock_Ticker']):
        #     if st.button(f"Remove {ticker}", key=f"remove_{i}"):
        #         st.session_state['portfolio'] = st.session_state['portfolio'].drop(index=i).reset_index(drop=True)

        # Display portfolio after removal
        st.write(st.session_state['portfolio'])

    #############################################################################################################################

    # Run Analysis if at least one stock is selected
    # Only can run 5 stocks at a time
    if st.session_state['portfolio'].shape[0] > 0:
        run_analysis = st.checkbox('Run Analysis')

        if run_analysis:
            # iterate through the values in the 'Stock_Ticker' column using iteritems()
            for index, value in st.session_state['portfolio']['Stock_Ticker'].iteritems():
                with st.expander(f"{value}"):
                    st.subheader(f'Processing - {value}:')

                    # RUN OPTIONS
                    # Option 1: If Stock is part of top 10, then just grab results straight from S3
                    # TOP 10 stocks in SP500 by index weight:
                    top_10_stocks = ['aapl', 'msft', 'amzn', 'nvda', 'googl', 'brk.b', 'goog', 'tsla', 'unh', 'meta']
                    if value in top_10_stocks:
                        df = get_analysis_data(value)
                        st.write(df.head(10))

                    # Option 2: If New Stock, then call new_stock_article_fetcher DAG
                    else:
                        # API Call to Airflow to execute process_audio_files_dag
                        data = {
                                "dag_run_id": "",
                                "conf": {"stock": value}
                                }
                        response = requests.post(url = f'http://{airflow_webserver}/api/v1/dags/new_stock_article_fetcher/dagRuns', json=data, auth=('airflow2','airflow2'))
                        if response.status_code == 409:
                            st.error(f'{value} data up-to-date in S3!')

                        dag_run_id = response.json()['dag_run_id']
                        st.write(f"Dag_run_id: {dag_run_id}")

                        starttime = time.time()
                        while check_dag_status("new_stock_article_fetcher") not in ('failed', 'success'):
                            time.sleep(10.0 - ((time.time() - starttime) % 10.0))

                        # IF DAG runs successfully then get the data
                        if check_dag_status("new_stock_article_fetcher") == 'success':
                            # get data from S3
                            df = get_analysis_data(value)
                            # Stop the script if there are no articles and display an error
                            if df.empty:
                                st.error(f'{value} has no articles on Seeking Alpha, try a different stock')
                                st.stop()
                            
                            st.write(df.head(10).sort_values('publish_date', ascending=False))

                    data2 = {'email': st.session_state.email}
                    headers = {"Authorization": f"Bearer {st.session_state.access_token}"}
                    res2 = requests.post(f'http://{fastapi_webserver}/update_api_calls', json=data2, headers = headers)
                    st.session_state['calls_left'] = st.session_state.calls_left - 1
                    # Summary of results
                    # group by 'Sentiment' and get counts
                    sentiment_counts = df.groupby(['sentiment']).size().reset_index(name='count').sort_values(by='count',
                                                                                                              ascending=False)
                    st.subheader(f'Article Summary for {value}')
                    st.write(sentiment_counts)

                    # Summarize the Positive and Negative articles
                    # st.write(df.loc[df['sentiment'] == 'positive', 'bart_summary'])
                    final_df = df.groupby(['sentiment'], as_index=False).agg({'bart_summary': ' '.join})
                    try:
                        pos_summary = final_df[final_df['sentiment'] == 'positive']['bart_summary'].item()
                    except:
                        pos_summary = ''
                    try:
                        neg_summary = final_df[final_df['sentiment'] == 'negative']['bart_summary'].item()
                    except:
                        neg_summary = ''

                    # Use Facebook BART model to summarize the aggregation of summaries
                    # IF there are no positive or negative articles
                    if pos_summary != '':
                        pos_overall_summary = article_summary('positive', pos_summary)
                    else:
                        pos_overall_summary = 'No positive sentiment articles'
                    if neg_summary != '':
                        neg_overall_summary = article_summary('negative', neg_summary)
                    else:
                        neg_overall_summary = 'No negative sentiment articles'
                    

                    # Display results
                    col1, col2 = st.columns(2)
                    with col1:
                        st.header("Positive Summary")
                        st.write(pos_overall_summary)

                    with col2:
                        st.header("Negative Summary")
                        st.write(neg_overall_summary)

                # ADD THE STOCKS TO SNOWFLAKE LOGS
                pos_overall_summary = pos_overall_summary.replace("'", "\\'")
                neg_overall_summary = neg_overall_summary.replace("'", "\\'")


                try:
                    pos_count = sentiment_counts.loc[sentiment_counts['sentiment'] == 'positive', 'count'].values[0]
                except:
                    pos_count = 0
                try:
                    neu_count = sentiment_counts.loc[sentiment_counts['sentiment'] == 'neutral', 'count'].values[0]
                except:
                    neu_count = 0
                try:
                    neg_count = sentiment_counts.loc[sentiment_counts['sentiment'] == 'negative', 'count'].values[0]
                except:
                    neg_count = 0

                db_util.add_stock_run(st.session_state.email, value,
                                      pos_count,
                                      neu_count,
                                      neg_count, pos_overall_summary, neg_overall_summary)

#############################################################################################################################


# # If the user has uploaded a file
# if file is not None:
#     # Read the Excel file into a Pandas DataFrame
#     df = pd.read_excel(file, sheet_name='Sheet1')
#     # Display the contents of the DataFrame
#     st.write(df)

# # Selecting Ticker for Analysis from the List Available
# add_item = st.text_input("**Add Ticker**")
# if add_item:
#     # Load the Excel file
#     workbook = openpyxl.load_workbook("C:/Users/rumij/Downloads/ticker.xlsx")
#     # Select the worksheet you want to add the item to
#     worksheet = workbook["Sheet1"]
#     # Add a new item to the worksheet
#     new_item = add_item
#     worksheet.append([new_item])
#     # Save the changes to the Excel file
#     workbook.save("C:/Users/rumij/Downloads/ticker.xlsx")

# remove_item = st.text_input("**Delete a Ticker**")
# if remove_item:
#     # Load the Excel file
#     workbook = openpyxl.load_workbook("C:/Users/rumij/Downloads/ticker.xlsx")
#     # Select the worksheet you want to add the item to
#     worksheet = workbook["Sheet1"]
#     # Add a new item to the worksheet
#     for row in worksheet.iter_rows():
#         for cell in row:
#             if cell.value == remove_item:
#             #     # Delete the row containing the cell
#                 worksheet.delete_rows(cell.row)
#                 # Save the changes to the Excel file
#                 workbook.save("C:/Users/rumij/Downloads/ticker.xlsx")
#                 break  # Exit the inner loop after deleting the row

# my_list = []
# stocks = pd.read_excel('C:/Users/rumij/Downloads/ticker.xlsx')
# stocks = stocks['Ticker'].unique()
# # selected_option = st.selectbox("**Select an option**", stocks)
# selected_option = st.multiselect("**Select Any 5 Tickers for Analysis**", stocks, max_selections=2)

# # my_list.append(selected_option)
# my_list = 'You selected: ' + ', '.join(selected_option)

# st.write(my_list)
